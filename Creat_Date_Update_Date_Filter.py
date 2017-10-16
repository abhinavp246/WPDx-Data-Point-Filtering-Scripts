''' This script will check the submission date of the data points, and the updated date. If the updated date 
occurs before the date of submission, the data point will be deleted. 
'''

#Load libraries and name variables.
from openpyxl import load_workbook, Workbook
wb = load_workbook('WPDxDateTime.xlsx') #Name of workbook you want to import.
ws = wb.active
wb1 = Workbook()
ws1 = wb1.create_sheet()
ws2 = wb1.create_sheet()
myList = []
#Traverse through the rows of the excel sheet.
for i in range (1, 351004): #for i in range (1, NUMBER OF ROWS IN YOUR SHEET + 1)
    d1 = ws.cell(row = i, column = 37) #column = Column # with the submission date.
    d2 = ws.cell(row = i, column = 40) #column = Column # with the updated date.
    if d2.value >= d1.value:
        for p in range (1,41): #for p in range (1, NUMBER OF COLUMNS IN YOUR SHEET + 1)
            store_val = ws.cell(row=i, column=p)
            myList.append(store_val.value)
        ws1.append(myList)
    myList = []
    if d2.value < d1.value:
        for p in range (1,41):
            store_val1 = ws.cell(row=i, column=p)
            myList.append(store_val1.value)
        ws2.append(myList)
    myList = []
wb1.save('AllDatesFiltered.xlsx') #Name of workbook being exported. 
