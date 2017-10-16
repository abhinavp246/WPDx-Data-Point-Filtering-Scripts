
''' This script will filter through any data points that were submitted in the future. NOTE: TO GET TODAY'S
DATE AS AN EXCEL INT, TYPE IT INTO AN EXCEL CELL IN THE MM/DD/YYYY FORMAT AND USE THE FORMULA =DATEVALUE(CELL).
'''

#Load all of the required libraries, and create a workbook.
from openpyxl import load_workbook, Workbook
wb = load_workbook('WPDxDateTime.xlsx') #THIS IS THE FILE NAME YOU WILL BE INPUTTING.
ws = wb.active
wb1 = Workbook()
ws1 = wb1.create_sheet()
ws2 = wb1.create_sheet()
ws3 = wb1.create_sheet()
myList = []
#Traverse through the rows of the WPDx excel sheet.
for i in range (1, 351004): #for i in range (1, NUMBER OF ROWS + 1)
    d1 = ws.cell(row = i, column = 37)
    if d2.value < 9999999 #REPLACE THIS NUMBER WITH TODAY'S DATE IN EXCEL FORMAT.
        for p in range (1,41):
            store_val = ws.cell(row=i, column=p)
            myList.append(store_val.value)
        ws1.append(myList)
    myList = []
    if d2.value > 9999999: #REPLACE THIS NUMBER WITH TODAY'S DATE IN EXCEL FORMAT.
        for p in range (1,41):
            store_val1 = ws.cell(row=i, column=p)
            myList.append(store_val1.value)
        ws2.append(myList) #Add it to the delete list if it's a future date.
    myList = []
wb1.save('AllDatesFiltered.xlsx') #THIS IS THE NAME OF THE OUTPUT FILE. CHANGE IT TO WHATEVER YOU WANT. 


