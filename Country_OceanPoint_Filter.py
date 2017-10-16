''' This script will eliminate mismatched country points nd ocean points. Make sure your excel 
    sheet has been formatted before using this script (there's a guide on how to do this in the same 
    directory as this file. 
'''

#Importing the required libraries, and creating new workbooks and variables.
from openpyxl import load_workbook, Workbook
wb = load_workbook('WPDx Validated copy.xlsx')
wb1 = Workbook()
ws1 = wb1.create_sheet()
ws2 = wb1.create_sheet()
ws = wb.active
myList = []
#Traversing through the rows, and assigning country names to d1 and d2.
for i in range (1,358752): #for i in range (1, TOTAL NUMBER OF ROWS + 1).
    d1 = ws.cell(row=i, column = 3)
    d2 = ws.cell(row=i, column = 30)
    if d1.value == d2.value: #If the two strings are they same (i.e they match), add them to the new workbook.
        for p in range(1,31): #for p in range (1, TOTAL NUMBER OF COLUMNS + 1)
            store_val = ws.cell(row = i, column = p)
            myList.append(store_val.value)
        ws1.append(myList)
    myList = []
    if d1.value != d2.value: #If the two strings are not the same, add them to the 'deleted points' sheet.
        for p in range(1,31):
            store_val1 = ws.cell(row= i, column = p)
            myList.append(store_val1.value)
        ws2.append(myList)
    myList = []
wb1.save('WPDxFinalFilter1.xlsx') #Save and exit.























